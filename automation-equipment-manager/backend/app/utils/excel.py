from datetime import date, datetime, time
from io import BytesIO

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app import models, schemas
from app.utils.enums import EQUIPMENT_STATUSES


EQUIPMENT_COLUMNS = [
    ("equipment_code", "设备编号"),
    ("equipment_name", "设备名称"),
    ("equipment_type", "设备类型"),
    ("brand", "品牌"),
    ("supplier", "供应商"),
    ("purchase_date", "购买日期"),
    ("purchase_price", "购买金额"),
    ("current_status", "当前状态"),
    ("current_location", "当前位置"),
    ("current_product_code", "当前生产货号"),
    ("manager", "负责人"),
    ("remark", "备注"),
]

EQUIPMENT_LEDGER_IMPORT_COLUMNS = [
    ("registration_code", "登记编号"),
    ("factory_area", "厂区"),
    ("department", "部门"),
    ("primary_category", "一级类别"),
    ("equipment_type", "二级类别"),
    ("equipment_name", "名称"),
    ("specification", "规格型号"),
    ("product_code", "货号"),
    ("unit", "单位"),
    ("purchase_date", "购买日期"),
    ("brand_supplier", "品牌厂家"),
    ("manager", "责任人"),
    ("usage_status", "使用状态"),
    ("current_location", "目前所在位置"),
    ("current_factory_area", "目前所在厂区"),
    ("remark", "备注"),
    ("registrar", "登记人"),
    ("created_at", "创建时间"),
    ("updated_at", "更新时间"),
]

EQUIPMENT_LEDGER_STATUS_MAP = {
    "正常使用": "生产中",
    "使用中": "生产中",
    "生产中": "生产中",
    "备用": "待用",
    "闲置": "待用",
    "待用": "待用",
    "维修": "维修中",
    "维修中": "维修中",
    "待维修": "维修中",
    "故障": "维修中",
    "异常": "维修中",
    "保养": "保养中",
    "保养中": "保养中",
    "调试": "调试中",
    "调试中": "调试中",
    "待验收": "待验收",
    "报废": "报废",
    "停用": "停用",
}


def normalize_equipment_status(usage_status, current_factory_area=None):
    factory_area = _normalize_text(current_factory_area)
    if factory_area and "".join(factory_area.split()) == "外厂":
        return "外发中"

    status = _normalize_text(usage_status)
    if not status:
        return "待用"

    status_key = "".join(status.split())
    return EQUIPMENT_LEDGER_STATUS_MAP.get(status_key, "待用")

OUTSOURCE_COLUMNS = [
    ("equipment_code", "设备编号"),
    ("equipment_name", "设备名称"),
    ("outsource_company", "外发单位"),
    ("contact_person", "联系人"),
    ("contact_phone", "联系电话"),
    ("outsource_reason", "外发原因"),
    ("outsource_date", "外发日期"),
    ("expected_return_date", "预计返回日期"),
    ("actual_return_date", "实际返回日期"),
    ("status", "状态"),
    ("return_status", "返回后状态"),
    ("operator", "操作人"),
    ("remark", "备注"),
]

PRODUCTION_COLUMNS = [
    ("equipment_code", "设备编号"),
    ("equipment_name", "设备名称"),
    ("product_code", "生产货号"),
    ("product_name", "产品名称"),
    ("department", "使用部门"),
    ("start_time", "开始时间"),
    ("end_time", "结束时间"),
    ("operator", "操作人员"),
    ("output_qty", "产量"),
    ("production_status", "生产状态"),
    ("remark", "备注"),
]

REPAIR_COLUMNS = [
    ("equipment_code", "设备编号"),
    ("equipment_name", "设备名称"),
    ("issue_time", "异常时间"),
    ("issue_description", "异常描述"),
    ("issue_level", "异常等级"),
    ("reporter", "反馈人"),
    ("handler", "处理人"),
    ("repair_status", "处理状态"),
    ("repair_method", "处理方法"),
    ("finish_time", "完成时间"),
    ("downtime_minutes", "停机分钟数"),
    ("remark", "备注"),
]

MAINTENANCE_COLUMNS = [
    ("equipment_code", "设备编号"),
    ("equipment_name", "设备名称"),
    ("maintenance_type", "保养类型"),
    ("maintenance_content", "保养内容"),
    ("plan_date", "计划日期"),
    ("actual_date", "实际日期"),
    ("maintainer", "保养人"),
    ("result", "保养结果"),
    ("next_date", "下次保养日期"),
    ("remark", "备注"),
]


def _date_to_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _normalize_date(value, row_number: int, column_name: str):
    if value in [None, ""]:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    raise HTTPException(status_code=400, detail=f"第 {row_number} 行 {column_name} 日期格式不正确，应为 YYYY-MM-DD")


def _normalize_date_or_none(value):
    if value in [None, ""]:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def _normalize_float(value, row_number: int, column_name: str):
    if value in [None, ""]:
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=f"第 {row_number} 行 {column_name} 必须是数字") from exc


def _normalize_int(value, row_number: int, column_name: str):
    if value in [None, ""]:
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=f"第 {row_number} 行 {column_name} 必须是整数") from exc


def _normalize_datetime(value, row_number: int, column_name: str):
    if value in [None, ""]:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, str):
        value = value.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y-%m-%dT%H:%M",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    raise HTTPException(status_code=400, detail=f"第 {row_number} 行 {column_name} 日期时间格式不正确，应为 YYYY-MM-DD HH:MM")


def _normalize_text(value):
    if value in [None, ""]:
        return None
    return str(value).strip()


def _required_text(value, row_number: int, column_name: str):
    text = _normalize_text(value)
    if not text:
        raise HTTPException(status_code=400, detail=f"第 {row_number} 行缺少必填字段：{column_name}")
    return text


def _workbook_to_response(workbook: Workbook) -> BytesIO:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _style_sheet(sheet):
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F9E6D")
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = 18


def _build_template(title: str, columns: list[tuple[str, str]], example: list, notes: list[str] | None = None) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append([label for _key, label in columns])
    sheet.append(example)
    if notes:
        sheet.append([])
        for note in notes:
            sheet.append([note])
    _style_sheet(sheet)
    return _workbook_to_response(workbook)


def _build_export(title: str, columns: list[tuple[str, str]], records: list[dict]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append([label for _key, label in columns])
    for record in records:
        sheet.append([record.get(key) for key, _label in columns])
    _style_sheet(sheet)
    return _workbook_to_response(workbook)


def _load_active_sheet(file_bytes: bytes):
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel 文件无法读取，请确认是 .xlsx 文件") from exc
    return workbook.active


def _read_rows(file_bytes: bytes, columns: list[tuple[str, str]], required_labels: list[str]):
    sheet = _load_active_sheet(file_bytes)
    headers = [cell.value for cell in sheet[1]]
    header_to_index = {label: index for index, label in enumerate(headers)}
    missing = [label for label in required_labels if label not in header_to_index]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必填列：{', '.join(missing)}")

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in [None, ""] for value in row):
            continue
        row_data = {}
        for key, label in columns:
            index = header_to_index.get(label)
            row_data[key] = row[index] if index is not None and index < len(row) else None
        yield row_number, row_data


def build_equipment_template() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "设备台账"
    sheet.append([label for _key, label in EQUIPMENT_COLUMNS])
    sheet.append(["AUTO-LSJ-001", "自动锁螺丝机001", "锁螺丝机", "德立宏", "德立宏供应商", "2026-04-24", 58000, "待用", "瑞海装配部一线", "HK-20260424-A01", "张三", "示例数据，导入前可删除"])
    sheet.append([])
    sheet.append(["说明：当前状态必须是以下固定选项之一："])
    sheet.append([", ".join(EQUIPMENT_STATUSES)])

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F9E6D")
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = 18
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["I"].width = 24
    sheet.column_dimensions["L"].width = 28
    return _workbook_to_response(workbook)


def build_equipment_export(equipment_list: list[models.Equipment]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "设备台账"
    sheet.append([label for _key, label in EQUIPMENT_COLUMNS])
    for equipment in equipment_list:
        sheet.append([
            equipment.equipment_code,
            equipment.equipment_name,
            equipment.equipment_type,
            equipment.brand,
            equipment.supplier,
            _date_to_text(equipment.purchase_date),
            equipment.purchase_price,
            equipment.current_status,
            equipment.current_location,
            equipment.current_product_code,
            equipment.manager,
            equipment.remark,
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F9E6D")
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = 18
    return _workbook_to_response(workbook)


def parse_equipment_import(file_bytes: bytes) -> list[schemas.EquipmentCreate]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel 文件无法读取，请确认是 .xlsx 文件") from exc

    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    header_to_index = {label: index for index, label in enumerate(headers)}
    missing = [label for _key, label in EQUIPMENT_COLUMNS[:3] if label not in header_to_index]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必填列：{', '.join(missing)}")

    records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in [None, ""] for value in row):
            continue
        row_data = {}
        for key, label in EQUIPMENT_COLUMNS:
            index = header_to_index.get(label)
            row_data[key] = row[index] if index is not None and index < len(row) else None

        row_data["purchase_date"] = _normalize_date(row_data["purchase_date"], row_number, "购买日期")
        row_data["purchase_price"] = _normalize_float(row_data["purchase_price"], row_number, "购买金额")
        row_data["current_status"] = row_data["current_status"] or "待用"

        try:
            records.append(schemas.EquipmentCreate(**row_data))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行数据校验失败：{exc}") from exc

    if not records:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的数据")
    return records


def parse_equipment_ledger_import(file_bytes: bytes):
    sheet = _load_active_sheet(file_bytes)
    headers = [_normalize_text(cell.value) for cell in sheet[1]]
    header_to_index = {label: index for index, label in enumerate(headers) if label}
    missing = [label for label in ["登记编号", "名称"] if label not in header_to_index]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必填列：{', '.join(missing)}")

    records = []
    failures = []
    total_count = 0
    skipped_count = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        total_count += 1
        if not any(value not in [None, ""] for value in row):
            skipped_count += 1
            continue

        row_data = {}
        for key, label in EQUIPMENT_LEDGER_IMPORT_COLUMNS:
            index = header_to_index.get(label)
            row_data[key] = row[index] if index is not None and index < len(row) else None

        equipment_code = _normalize_text(row_data["registration_code"])
        equipment_name = _normalize_text(row_data["equipment_name"])
        if not equipment_code:
            failures.append({"row_number": row_number, "reason": "登记编号不能为空"})
            continue
        if not equipment_name:
            failures.append({"row_number": row_number, "reason": "名称不能为空"})
            continue

        current_status = normalize_equipment_status(row_data["usage_status"], row_data["current_factory_area"])
        equipment_type = _normalize_text(row_data["equipment_type"]) or _normalize_text(row_data["primary_category"]) or "未分类"
        brand_supplier = _normalize_text(row_data["brand_supplier"])

        payload = {
            "equipment_code": equipment_code,
            "equipment_name": equipment_name,
            "equipment_type": equipment_type,
            "brand": brand_supplier,
            "supplier": brand_supplier,
            "purchase_date": _normalize_date_or_none(row_data["purchase_date"]),
            "current_status": current_status,
            "current_location": _normalize_text(row_data["current_location"]),
            "current_product_code": _normalize_text(row_data["product_code"]),
            "manager": _normalize_text(row_data["manager"]),
            "remark": _normalize_text(row_data["remark"]),
        }

        try:
            records.append({"row_number": row_number, "equipment": schemas.EquipmentCreate(**payload)})
        except Exception as exc:
            failures.append({"row_number": row_number, "reason": f"数据校验失败：{exc}"})

    return {
        "records": records,
        "failures": failures,
        "total_count": total_count,
        "skipped_count": skipped_count,
    }


def build_outsource_template() -> BytesIO:
    return _build_template(
        "外发记录",
        OUTSOURCE_COLUMNS,
        ["AUTO-LSJ-001", "自动锁螺丝机001", "外协维修厂", "李四", "13800000000", "电控维修", "2026-04-24", "2026-04-30", "", "外发中", "待用", "张三", "示例数据，导入前可删除"],
        ["说明：设备编号、外发单位、外发日期、预计返回日期为必填。", "如果填写实际返回日期，系统会自动登记返回；返回后状态可填：待用、调试中、生产中。"],
    )


def build_outsource_export(records: list[dict]) -> BytesIO:
    export_records = []
    for record in records:
        export_records.append({
            **record,
            "outsource_date": _date_to_text(record.get("outsource_date")),
            "expected_return_date": _date_to_text(record.get("expected_return_date")),
            "actual_return_date": _date_to_text(record.get("actual_return_date")),
            "return_status": "",
        })
    return _build_export("外发记录", OUTSOURCE_COLUMNS, export_records)


def parse_outsource_import(file_bytes: bytes) -> list[tuple[str, schemas.EquipmentOutsourceCreate, date | None, str]]:
    records = []
    for row_number, row_data in _read_rows(file_bytes, OUTSOURCE_COLUMNS, ["设备编号", "外发单位", "外发日期", "预计返回日期"]):
        equipment_code = _required_text(row_data["equipment_code"], row_number, "设备编号")
        actual_return_date = _normalize_date(row_data["actual_return_date"], row_number, "实际返回日期")
        return_status = _normalize_text(row_data["return_status"]) or "待用"
        payload = {
            "outsource_company": _required_text(row_data["outsource_company"], row_number, "外发单位"),
            "contact_person": _normalize_text(row_data["contact_person"]),
            "contact_phone": _normalize_text(row_data["contact_phone"]),
            "outsource_reason": _normalize_text(row_data["outsource_reason"]),
            "outsource_date": _normalize_date(row_data["outsource_date"], row_number, "外发日期"),
            "expected_return_date": _normalize_date(row_data["expected_return_date"], row_number, "预计返回日期"),
            "operator": _normalize_text(row_data["operator"]),
            "remark": _normalize_text(row_data["remark"]),
        }
        try:
            records.append((equipment_code, schemas.EquipmentOutsourceCreate(**payload), actual_return_date, return_status))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行数据校验失败：{exc}") from exc

    if not records:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的数据")
    return records


def build_production_template() -> BytesIO:
    return _build_template(
        "生产记录",
        PRODUCTION_COLUMNS,
        ["AUTO-LSJ-001", "自动锁螺丝机001", "HK-20260424-A01", "壳体组件", "瑞海装配部", "2026-04-24 08:00", "", "王五", 120, "生产中", "示例数据，导入前可删除"],
        ["说明：设备编号、生产货号为必填；开始时间为空时会使用导入时刻。"],
    )


def build_production_export(records: list[dict]) -> BytesIO:
    export_records = []
    for record in records:
        export_records.append({
            **record,
            "start_time": _date_to_text(record.get("start_time")),
            "end_time": _date_to_text(record.get("end_time")),
        })
    return _build_export("生产记录", PRODUCTION_COLUMNS, export_records)


def parse_production_import(file_bytes: bytes) -> list[tuple[str, schemas.EquipmentProductionCreate]]:
    records = []
    for row_number, row_data in _read_rows(file_bytes, PRODUCTION_COLUMNS, ["设备编号", "生产货号"]):
        equipment_code = _required_text(row_data["equipment_code"], row_number, "设备编号")
        payload = {
            "product_code": _required_text(row_data["product_code"], row_number, "生产货号"),
            "product_name": _normalize_text(row_data["product_name"]),
            "department": _normalize_text(row_data["department"]),
            "start_time": _normalize_datetime(row_data["start_time"], row_number, "开始时间"),
            "end_time": _normalize_datetime(row_data["end_time"], row_number, "结束时间"),
            "operator": _normalize_text(row_data["operator"]),
            "output_qty": _normalize_int(row_data["output_qty"], row_number, "产量"),
            "production_status": _normalize_text(row_data["production_status"]),
            "remark": _normalize_text(row_data["remark"]),
        }
        try:
            records.append((equipment_code, schemas.EquipmentProductionCreate(**payload)))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行数据校验失败：{exc}") from exc

    if not records:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的数据")
    return records


def build_repair_template() -> BytesIO:
    return _build_template(
        "维修异常",
        REPAIR_COLUMNS,
        ["AUTO-LSJ-001", "自动锁螺丝机001", "2026-04-24 09:30", "夹具定位异常", "一般", "赵六", "钱七", "待处理", "检查定位传感器", "", 30, "示例数据，导入前可删除"],
        ["说明：设备编号、异常描述为必填；异常等级可填：轻微、一般、严重、重大；处理状态可填：待处理、处理中、已解决、需外发。"],
    )


def build_repair_export(records: list[dict]) -> BytesIO:
    export_records = []
    for record in records:
        export_records.append({
            **record,
            "issue_time": _date_to_text(record.get("issue_time")),
            "finish_time": _date_to_text(record.get("finish_time")),
        })
    return _build_export("维修异常", REPAIR_COLUMNS, export_records)


def parse_repair_import(file_bytes: bytes) -> list[tuple[str, schemas.EquipmentRepairCreate]]:
    records = []
    for row_number, row_data in _read_rows(file_bytes, REPAIR_COLUMNS, ["设备编号", "异常描述"]):
        equipment_code = _required_text(row_data["equipment_code"], row_number, "设备编号")
        payload = {
            "issue_time": _normalize_datetime(row_data["issue_time"], row_number, "异常时间"),
            "issue_description": _required_text(row_data["issue_description"], row_number, "异常描述"),
            "issue_level": _normalize_text(row_data["issue_level"]) or "一般",
            "reporter": _normalize_text(row_data["reporter"]),
            "handler": _normalize_text(row_data["handler"]),
            "repair_status": _normalize_text(row_data["repair_status"]) or "待处理",
            "repair_method": _normalize_text(row_data["repair_method"]),
            "finish_time": _normalize_datetime(row_data["finish_time"], row_number, "完成时间"),
            "downtime_minutes": _normalize_int(row_data["downtime_minutes"], row_number, "停机分钟数"),
            "remark": _normalize_text(row_data["remark"]),
        }
        try:
            records.append((equipment_code, schemas.EquipmentRepairCreate(**payload)))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行数据校验失败：{exc}") from exc

    if not records:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的数据")
    return records


def build_maintenance_template() -> BytesIO:
    return _build_template(
        "保养记录",
        MAINTENANCE_COLUMNS,
        ["AUTO-LSJ-001", "自动锁螺丝机001", "日常保养", "清洁导轨并检查气路", "2026-04-24", "2026-04-24", "孙八", "正常", "2026-05-24", "示例数据，导入前可删除"],
        ["说明：设备编号、保养内容为必填；保养类型可填：日常保养、周保养、月保养、年度保养、临时保养。"],
    )


def build_maintenance_export(records: list[dict]) -> BytesIO:
    export_records = []
    for record in records:
        export_records.append({
            **record,
            "plan_date": _date_to_text(record.get("plan_date")),
            "actual_date": _date_to_text(record.get("actual_date")),
            "next_date": _date_to_text(record.get("next_date")),
        })
    return _build_export("保养记录", MAINTENANCE_COLUMNS, export_records)


def parse_maintenance_import(file_bytes: bytes) -> list[tuple[str, schemas.EquipmentMaintenanceCreate]]:
    records = []
    for row_number, row_data in _read_rows(file_bytes, MAINTENANCE_COLUMNS, ["设备编号", "保养内容"]):
        equipment_code = _required_text(row_data["equipment_code"], row_number, "设备编号")
        payload = {
            "maintenance_type": _normalize_text(row_data["maintenance_type"]) or "日常保养",
            "maintenance_content": _required_text(row_data["maintenance_content"], row_number, "保养内容"),
            "plan_date": _normalize_date(row_data["plan_date"], row_number, "计划日期"),
            "actual_date": _normalize_date(row_data["actual_date"], row_number, "实际日期"),
            "maintainer": _normalize_text(row_data["maintainer"]),
            "result": _normalize_text(row_data["result"]),
            "next_date": _normalize_date(row_data["next_date"], row_number, "下次保养日期"),
            "remark": _normalize_text(row_data["remark"]),
        }
        try:
            records.append((equipment_code, schemas.EquipmentMaintenanceCreate(**payload)))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行数据校验失败：{exc}") from exc

    if not records:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的数据")
    return records
